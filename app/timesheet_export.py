"""Экспорт графика в Excel в виде табеля (формат, близкий к типовой таблице учёта времени)."""

from __future__ import annotations

import calendar
import io
from datetime import date
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MONTHS_RU_UPPER = (
    "",
    "ЯНВАРЬ",
    "ФЕВРАЛЬ",
    "МАРТ",
    "АПРЕЛЬ",
    "МАЙ",
    "ИЮНЬ",
    "ИЮЛЬ",
    "АВГУСТ",
    "СЕНТЯБРЬ",
    "ОКТЯБРЬ",
    "НОЯБРЬ",
    "ДЕКАБРЬ",
)

WD_SHORT_RU = ("пн", "вт", "ср", "чт", "пт", "сб", "вс")

FILL_HEADER = PatternFill(start_color="FFE8EEF5", end_color="FFE8EEF5", fill_type="solid")
FILL_WEEKEND_HEAD = PatternFill(start_color="FFDCFCE7", end_color="FFDCFCE7", fill_type="solid")
FILL_HOLIDAY_HEAD = PatternFill(start_color="FFFFEDD5", end_color="FFFFEDD5", fill_type="solid")
FILL_DAYOFF = PatternFill(start_color="FFBBF7D0", end_color="FFBBF7D0", fill_type="solid")
FILL_WEEKEND_WORK = PatternFill(start_color="FFC7D2FE", end_color="FFC7D2FE", fill_type="solid")
FILL_EARLY = PatternFill(start_color="FFE9D5FF", end_color="FFE9D5FF", fill_type="solid")
FILL_LATE = PatternFill(start_color="FFBFDBFE", end_color="FFBFDBFE", fill_type="solid")

FONT_TITLE = Font(name="Calibri", size=14, bold=True)
FONT_SUB = Font(name="Calibri", size=11, bold=True)
FONT_HEADER = Font(name="Calibri", size=9, bold=True)
FONT_CELL = Font(name="Calibri", size=9)
FONT_SMALL = Font(name="Calibri", size=8, italic=True, color="FF5A6578")

THIN = Side(style="thin", color="FFCCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)


def _day_is_rest(d: date, holidays: set[date]) -> bool:
    if d in holidays:
        return True
    return d.weekday() >= 5


def _fmt_hours(value: float) -> str:
    if value <= 0:
        return ""
    r = round(value, 2)
    if abs(r - int(r)) < 0.05:
        return str(int(round(r)))
    return f"{r:.1f}".rstrip("0").rstrip(".")


def build_timesheet_excel_bytes(payload: dict[str, Any]) -> tuple[bytes, str]:
    """
    payload: year, month, holidays (list iso), rows (assignments with date, employee_name, duration_hours),
    people (sorted list: name, rate, preferred_hours, norm_hours_per_workday, position_label).
    """
    year = int(payload["year"])
    month = int(payload["month"])
    holidays = {date.fromisoformat(x) for x in payload.get("holidays", []) if x}
    rows_raw = payload.get("rows", [])
    people = payload.get("people", [])

    _, last_day = calendar.monthrange(year, month)
    month_days = [date(year, month, d) for d in range(1, last_day + 1)]

    hours_grid: dict[str, dict[date, float]] = {}
    for row in rows_raw:
        name = (row.get("employee_name") or "").strip()
        if not name:
            continue
        try:
            d = date.fromisoformat(str(row.get("date", "")))
        except ValueError:
            continue
        if d.year != year or d.month != month:
            continue
        dur = float(row.get("duration_hours") or 0)
        if dur <= 0:
            continue
        if name not in hours_grid:
            hours_grid[name] = {}
        hours_grid[name][d] = hours_grid[name].get(d, 0.0) + dur

    wb = Workbook()
    ws = wb.active
    ws.title = "Табель"

    meta_cols = 6
    first_day_col = meta_cols + 1
    last_day_col = first_day_col + len(month_days) - 1
    norm_col = last_day_col + 1
    fact_col = last_day_col + 2
    delta_col = last_day_col + 3
    carry_col = last_day_col + 4
    last_col = carry_col

    total_cols = last_col

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c1 = ws.cell(row=1, column=1, value="ТАБЕЛЬ учета использования рабочего времени")
    c1.font = FONT_TITLE
    c1.alignment = ALIGN_CENTER

    month_label = f"{MONTHS_RU_UPPER[month]} {year}"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    c2 = ws.cell(row=2, column=1, value=month_label)
    c2.font = FONT_SUB
    c2.alignment = ALIGN_CENTER

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_cols)
    legend = (
        "Условные обозначения: зелёный фон — выходной/праздник без часов (В); "
        "синий оттенок — отработка в выходной/праздник; фиолет./голуб. — начало с 8:00 / окончание в 22:00."
    )
    c3 = ws.cell(row=3, column=1, value=legend)
    c3.font = FONT_SMALL
    c3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    header_row = 5
    meta_headers = (
        "№ п/п",
        "Ф.И.О.",
        "Смена по графику",
        "Ставка",
        "Норма часов в день",
        "Должность",
    )
    for col, text in enumerate(meta_headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.border = BORDER
        cell.alignment = ALIGN_CENTER

    for i, d in enumerate(month_days):
        col = first_day_col + i
        wd = WD_SHORT_RU[d.weekday()]
        cell = ws.cell(row=header_row, column=col, value=f"{wd}\n{d.day}")
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER
        if d in holidays:
            cell.fill = FILL_HOLIDAY_HEAD
        elif d.weekday() >= 5:
            cell.fill = FILL_WEEKEND_HEAD
        else:
            cell.fill = FILL_HEADER

    for col, text in (
        (norm_col, "НОРМА"),
        (fact_col, "ФАКТ"),
        (delta_col, "Отклонение\n(факт − норма)"),
        (carry_col, "перенос часов\nна след. месяц"),
    ):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.border = BORDER
        cell.alignment = ALIGN_CENTER

    data_start = header_row + 1
    for idx, person in enumerate(people, start=1):
        r = data_start + idx - 1
        name = str(person.get("name", ""))
        rate = float(person.get("rate", 1.0) or 1.0)
        preferred = float(person.get("preferred_hours", 0.0) or 0.0)
        norm_day = float(person.get("norm_hours_per_workday", 0.0) or 0.0)
        position = str(person.get("position_label", "") or "")
        assigned = float(person.get("assigned_hours", 0.0) or 0.0)

        meta_values = (
            idx,
            name,
            "",
            rate,
            norm_day if norm_day > 0 else "",
            position,
        )
        for col, val in enumerate(meta_values, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = FONT_CELL
            cell.border = BORDER
            cell.alignment = ALIGN_LEFT if col == 2 else ALIGN_CENTER

        day_hours = hours_grid.get(name, {})
        for i, d in enumerate(month_days):
            col = first_day_col + i
            h = day_hours.get(d, 0.0)
            rest = _day_is_rest(d, holidays)
            cell = ws.cell(row=r, column=col)
            cell.border = BORDER
            cell.font = FONT_CELL
            cell.alignment = ALIGN_CENTER
            if h > 0:
                cell.value = _fmt_hours(h)
                starts: list[str] = []
                ends: list[str] = []
                for row in rows_raw:
                    if (row.get("employee_name") or "").strip() != name:
                        continue
                    try:
                        rd = date.fromisoformat(str(row.get("date", "")))
                    except ValueError:
                        continue
                    if rd != d:
                        continue
                    st = str(row.get("start_time") or "")
                    et = str(row.get("end_time") or "")
                    if st:
                        starts.append(st)
                    if et:
                        ends.append(et)
                start_t = min(starts) if starts else ""
                end_t = max(ends) if ends else ""
                if rest:
                    cell.fill = FILL_WEEKEND_WORK
                elif start_t and start_t <= "08:00":
                    cell.fill = FILL_EARLY
                elif end_t and end_t >= "22:00":
                    cell.fill = FILL_LATE
            elif rest:
                cell.value = "В"
                cell.fill = FILL_DAYOFF
            else:
                cell.value = ""

        delta = round(assigned - preferred, 1)
        for col, val in (
            (norm_col, round(preferred, 1)),
            (fact_col, round(assigned, 1)),
            (delta_col, delta),
            (carry_col, ""),
        ):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = FONT_CELL
            cell.border = BORDER
            cell.alignment = ALIGN_CENTER

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 18
    for i in range(len(month_days)):
        ws.column_dimensions[get_column_letter(first_day_col + i)].width = 5.5
    ws.column_dimensions[get_column_letter(norm_col)].width = 9
    ws.column_dimensions[get_column_letter(fact_col)].width = 9
    ws.column_dimensions[get_column_letter(delta_col)].width = 12
    ws.column_dimensions[get_column_letter(carry_col)].width = 14

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[3].height = 28
    ws.row_dimensions[header_row].height = 30

    freeze = f"{get_column_letter(first_day_col)}{data_start}"
    ws.freeze_panes = freeze

    buf = io.BytesIO()
    wb.save(buf)
    filename = f"ekc_tabel_{year}_{month:02d}.xlsx"
    return buf.getvalue(), filename
